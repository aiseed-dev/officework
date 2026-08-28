"""セルの中の飾り(richtext)が、触らなければ残るか。

    .venv/bin/python test/basic_rich.py

**模型には持ちません**(2026-08-28 発注者「CellRichText を共通ライブラリー
に入れないで」)。原本の `<r>` の並びをそのまま返す形なので、字を
書き替えたセルは普通の字になります。それが正しい振る舞いです。

Excel は飾りを共有文字列に、openpyxl はセルの中に書きます。どちらも
拾えるかを見ます。
"""
import os
import sys
import tempfile
import zipfile

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from officework import sheet

warui = 0


def check(cond, msg):
    global warui
    print(("  OK  " if cond else "× ") + msg)
    if not cond:
        warui += 1


def kazari(path, part="xl/sharedStrings.xml"):
    z = zipfile.ZipFile(path)
    if part not in z.namelist():
        return False
    x = z.read(part).decode("utf-8")
    return "<r>" in x or "<r " in x


tmp = tempfile.mkdtemp(prefix="rich-")
moto = os.path.join(tmp, "moto.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = CellRichText(
    [TextBlock(InlineFont(b=True, color="C00000"), "赤い太字"), "と普通の字"]
)
ws["A2"] = "触らない字"
ws["A3"] = CellRichText([TextBlock(InlineFont(i=True), "斜め"), "の字"])
wb.save(moto)

# ① 触らずに開いて保存すると、飾りが残る
b = sheet.Book.open(moto)
out1 = os.path.join(tmp, "sonomama.xlsx")
b.save(out1)
check(kazari(out1), "触らずに保存して飾りが残る")
check(openpyxl.load_workbook(out1)["Sheet"]["A1"].value == "赤い太字と普通の字",
      "字が読める")

# ② 別のセルを触っても、飾りのセルは残る
b2 = sheet.Book.open(moto)
b2[0]["B1"] = "後から足した字"
out2 = os.path.join(tmp, "tonari.xlsx")
b2.save(out2)
check(kazari(out2), "隣を触っても飾りが残る")

# ③ 飾りのセルの字を書き替えたら、普通の字になる(模型に持たないので)
b3 = sheet.Book.open(moto)
b3[0]["A1"] = "書き替えた字"
out3 = os.path.join(tmp, "kaeta.xlsx")
b3.save(out3)
z = zipfile.ZipFile(out3).read("xl/sharedStrings.xml").decode("utf-8")
check("書き替えた字" in z or "&#26360;" in z, "書き替えた字が入っている")
check(openpyxl.load_workbook(out3)["Sheet"]["A1"].value == "書き替えた字",
      "書き替えた字が読める")
# A3 は触っていないので飾りが残る
check("<r>" in z, "触っていない方(A3)の飾りは残る")

print("OK" if warui == 0 else "{} 件おかしい".format(warui))
sys.exit(1 if warui else 0)
