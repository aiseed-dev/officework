#!/usr/bin/env python3
"""表の基本のテスト — 毎日する操作だけ。合わなければその場で止まる。

    .venv/bin/python test/basic_xlsx.py

見るのは基本だけです: 値を書いて読む・式の計算結果・保存して開いて同じ・
行の挿入で式が追う・別シート参照・日付・結合と列幅の往復・openpyxl の目で読み返す。
"""
import datetime as dt
import tempfile
from pathlib import Path

from officework import sheet

tmp = Path(tempfile.mkdtemp(prefix="basic-xlsx-"))


BAD = []

def check(label, got, want):
    if got != want:
        BAD.append(label)
        print(f"× {label}: {want!r} のはずが {got!r}")
    else:
        print(f"  {label}: OK")


# 1. 値を書いて、そのまま読める
b = sheet.Book()
ws = b[0]
ws.cell(1, 1).value = "品名"
ws.cell(1, 2).value = 100
ws.cell(1, 3).value = 1.5
ws.cell(1, 4).value = dt.date(2026, 8, 26)
ws.cell(1, 5).value = True
check("字", ws.cell(1, 1).value, "品名")
check("整数", ws.cell(1, 2).value, 100)
check("小数", ws.cell(1, 3).value, 1.5)
check("日付", ws.cell(1, 4).value, dt.date(2026, 8, 26))
check("真偽", ws.cell(1, 5).value, True)

# 2. 式の計算結果が合う(四則・SUM・IF・AVERAGE・COUNT・文字列)
b = sheet.Book()
ws = b[0]
for i, v in enumerate([10, 20, 30, 40], start=1):
    ws.cell(i, 1).value = v
ws.cell(1, 2).value = "=A1+A2"
ws.cell(2, 2).value = "=A4-A1"
ws.cell(3, 2).value = "=A2*A3"
ws.cell(4, 2).value = "=A4/A2"
ws.cell(5, 2).value = "=SUM(A1:A4)"
ws.cell(6, 2).value = "=AVERAGE(A1:A4)"
ws.cell(7, 2).value = "=COUNT(A1:A4)"
ws.cell(8, 2).value = '=IF(A1<A2, "小さい", "大きくない")'
ws.cell(9, 2).value = "=MAX(A1:A4)-MIN(A1:A4)"
ws.cell(10, 2).value = '=A1&"円"'
b.recalc()
check("足し算", ws.cell(1, 2).value, 30)
check("引き算", ws.cell(2, 2).value, 30)
check("掛け算", ws.cell(3, 2).value, 600)
check("割り算", ws.cell(4, 2).value, 2)
check("SUM", ws.cell(5, 2).value, 100)
check("AVERAGE", ws.cell(6, 2).value, 25)
check("COUNT", ws.cell(7, 2).value, 4)
check("IF", ws.cell(8, 2).value, "小さい")
check("MAX-MIN", ws.cell(9, 2).value, 30)
check("文字の連結", ws.cell(10, 2).value, "10円")

# 3. 保存して開いて、値も式も同じ
p = str(tmp / "basic.xlsx")
b.save(p)
b2 = sheet.Book.open(p)
ws2 = b2[0]
check("開き直した値", ws2.cell(5, 2).value, 100)
check("開き直した式", ws2.formula("B5"), "=SUM(A1:A4)")

# 4. 行を挿すと式が追う
ws2.insert_row(1)
b2.recalc()
check("挿入後の式", ws2.formula("B6"), "=SUM(A2:A5)")
check("挿入後の値", ws2.cell(6, 2).value, 100)

# 5. 別シート参照
s2 = b2.create_sheet("集計")
s2.cell(1, 1).value = "=SUM(Sheet1!A2:A5)*2"
b2.recalc()
check("別シート参照", s2.cell(1, 1).value, 200)

# 6. 結合・列幅・書式が往復で残る
b3 = sheet.Book()
ws3 = b3[0]
ws3.merge_cells("A1:C1")
ws3.cell(1, 1).value = "題"
ws3.column_dimensions["B"].width = 25
ws3.cell(2, 1).value = 1234.5
ws3.cell(2, 1).number_format = "#,##0.0"
p3 = str(tmp / "fmt.xlsx")
b3.save(p3)
b4 = sheet.Book.open(p3)
check("結合の往復", ("A1:C1" in [str(m) for m in b4[0].merged_cell_ranges]), True)
check("列幅の往復", round(b4[0].column_dimensions["B"].width), 25)
check("表示形式の往復", b4[0].cell(2, 1).number_format, "#,##0.0")

# 7. 本家の目 — openpyxl で読み返して同じ
import openpyxl
wb = openpyxl.load_workbook(p, data_only=True)
check("openpyxl の目(値)", wb["Sheet1"]["B5"].value, 100)
wb2 = openpyxl.load_workbook(p)
check("openpyxl の目(式)", wb2["Sheet1"]["B5"].value, "=SUM(A1:A4)")
wb3 = openpyxl.load_workbook(p3)
check("openpyxl の目(結合)", "A1:C1" in [str(m) for m in wb3.active.merged_cells.ranges], True)
check("openpyxl の目(表示形式)", wb3.active["A2"].number_format, "#,##0.0")

if BAD:
    raise SystemExit(f"合わない物 {len(BAD)} 件: " + "、".join(BAD))
print("基本のテスト: 全部合った")
