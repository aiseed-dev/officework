"""ピボットが、アプリ無しで置けるか。

    .venv/bin/python test/basic_pivot.py

2026-08-29 発注者「ピボットの処理は polars をつかって」。
いままでは Python の polars を別プロセスで呼んでいたので、動いている
calc からしか使えませんでした。Rust の polars に移して、
`pip install officework` した人からも使えます。
"""
import os
import sys
import tempfile

import openpyxl

from officework import sheet

warui = 0


def check(cond, msg):
    global warui
    print(("  OK  " if cond else "× ") + msg)
    if not cond:
        warui += 1


def moto():
    b = sheet.Book()
    ws = b[0]
    data = [
        ["支店", "月", "金額"],
        ["東京", "4月", 1000], ["東京", "4月", 500], ["東京", "5月", 1200],
        ["大阪", "4月", 800], ["大阪", "5月", 300], ["大阪", "5月", 700],
    ]
    for r, row in enumerate(data, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(r, c).value = v
    return b, ws


# ① 行だけ(支店ごとの合計)
b, ws = moto()
h, w = ws.add_pivot("A1:C7", "E1", rows=["支店"], value="金額")
check((h, w) == (4, 2), "広さ: {} 行 × {} 列".format(h, w))
check(ws["F2"].value == 1800, "大阪の合計: {}".format(ws["F2"].value))
check(ws["F3"].value == 2700, "東京の合計: {}".format(ws["F3"].value))
check(ws["F4"].value == 4500, "総計: {}".format(ws["F4"].value))
check(ws["E1"].font.bold, "見出しが太字")
check(ws["E4"].font.bold, "総計が太字")

# ② 行×列(支店×月)。**列に広げると見出しが2行**になります
#    (1行目に「何を集計したか」と「どの見出しで広げたか」、2行目に列の名前)
b, ws = moto()
h, w = ws.add_pivot("A1:C7", "E1", rows=["支店"], cols=["月"], value="金額")
check((h, w) == (5, 4), "広さ: {} 行 × {} 列".format(h, w))
check([ws.cell(1, 5 + c).value for c in range(2)] == ["sum / 金額", "月"], "札の行")
check([ws.cell(2, 5 + c).value for c in range(4)] == ["支店", "4月", "5月", "Grand totals"],
      "見出し")
check(ws["F4"].value == 1500, "東京の4月: {}".format(ws["F4"].value))

# ③ 集計の仕方
# 東京は 1000 / 500 / 1200 → 件数3・平均900・最大1200・最小500
for agg, machi in [("count", 3), ("mean", 900), ("max", 1200), ("min", 500)]:
    b, ws = moto()
    ws.add_pivot("A1:C7", "E1", rows=["支店"], value="金額", agg=agg, totals=False)
    got = ws["F3"].value
    check(abs(got - machi) < 0.51, "{}: {}(待っていたのは {})".format(agg, got, machi))

# ④ 知らない集計の仕方は正直に断る
b, ws = moto()
try:
    ws.add_pivot("A1:C7", "E1", rows=["支店"], value="金額", agg="なんとか")
    check(False, "知らない集計を黙って受けた")
except Exception as e:
    check("なんとか" in str(e), "断りの文に名前が出る")

# ⑤ 保存して、本家の目でも読める
b, ws = moto()
ws.add_pivot("A1:C7", "E1", rows=["支店"], cols=["月"], value="金額")
out = os.path.join(tempfile.mkdtemp(), "pivot.xlsx")
b.save(out)
r = openpyxl.load_workbook(out)["Sheet1"]
check(r["H5"].value == 4500, "保存して開き直した総計: {}".format(r["H5"].value))

# ⑥ 小計と、日本語の札
b, ws = moto()
ws.cell(1, 3).value = "金額"
h, w = ws.add_pivot("A1:C7", "E1", rows=["支店", "月"], value="金額",
                    subtotals=True, grand_label="総計", subtotal_label="{} 小計")
mita = [[ws.cell(r, 5 + c).value for c in range(w)] for r in range(1, h + 1)]
check(["大阪 小計", None, 1800] in mita, "小計の行: {}".format(mita))
check(["総計", None, 4500] in mita, "総計の札が日本語")

print("OK" if warui == 0 else "{} 件おかしい".format(warui))
sys.exit(1 if warui else 0)
