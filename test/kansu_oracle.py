#!/usr/bin/env python3
"""関数の答えの表を作る。

kansu_cases.tsv の式を openpyxl で xlsx に並べ、LibreOffice を画面なしで
動かして計算させ、答えを kansu_kotae.tsv に書く。

答えの正は「本家が実際に計算した値」で、人も AI もここに手を入れない。
いまの本家は LibreOffice。M365 が使えるようになったら、同じ xlsx を
Excel で開いて保存し、--kotae にそのファイルを渡せば Excel の答えに
差し替わる(道具は同じ、正だけ替える)。

使い方:
    python3 test/kansu_oracle.py            # 生成 → 計算 → kansu_kotae.tsv
    python3 test/kansu_oracle.py --kotae 計算済み.xlsx   # 答えだけ吸い直す
"""

import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
CASES = HERE / "kansu_cases.tsv"
KOTAE = HERE / "kansu_kotae.tsv"


def load_cases():
    out = []
    for line in CASES.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        name, formula = line.split("\t", 1)
        out.append((name.strip(), formula.strip()))
    return out


def build_xlsx(cases, path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "kotae"
    for i, (name, formula) in enumerate(cases, start=1):
        ws.cell(row=i, column=1, value=name)
        # 式の字そのもの(答えと並べて読むため)
        ws.cell(row=i, column=2, value="'" + formula)
        ws.cell(row=i, column=3, value=formula)
    wb.save(path)


def compute(src, outdir):
    # LibreOffice の既定は「xlsx の式を読み直しても再計算しない」。
    # 専用のプロファイルに「常に再計算」を書いてから動かす
    prof = outdir / "profile"
    (prof / "user").mkdir(parents=True, exist_ok=True)
    (prof / "user/registrymodifications.xcu").write_text(
        '<?xml version="1.0"?>\n'
        '<oor:items xmlns:oor="http://openoffice.org/2001/registry">\n'
        ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
        '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
        ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
        '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
        "</oor:items>\n",
        encoding="utf-8",
    )
    # 出力は別の入れ物に(同じ場所だと入力と同じ名前になり、変換されない)
    saki = outdir / "out"
    saki.mkdir(exist_ok=True)
    r = subprocess.run(
        [
            "soffice",
            f"-env:UserInstallation=file://{prof}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(saki),
            str(src),
        ],
        capture_output=True,
        text=True,
        timeout=300,
    )
    out = saki / (Path(src).stem + ".xlsx")
    if not out.exists():
        raise SystemExit(f"LibreOffice が書き出せませんでした: {r.stderr}")
    return out


def suck(path, cases):
    import datetime

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for i, (name, formula) in enumerate(cases, start=1):
        v = ws.cell(row=i, column=3).value
        if v is None:
            v = ""
        elif isinstance(v, bool):
            v = "TRUE" if v else "FALSE"
        elif isinstance(v, datetime.datetime):
            # 日付の表示形式が付くと openpyxl は datetime で返す。
            # 中身は通し番号(1899-12-30 起点)なので、数に戻して書く
            delta = v - datetime.datetime(1899, 12, 30)
            v = repr(delta.days + delta.seconds / 86400)
        elif isinstance(v, datetime.time):
            # 時刻の表示形式も同じ — 日の割合に戻す
            v = repr((v.hour * 3600 + v.minute * 60 + v.second) / 86400)
        elif isinstance(v, float):
            v = repr(v)
        rows.append(f"{name}\t{formula}\t{v}")
    return rows


def main():
    cases = load_cases()
    if len(sys.argv) > 2 and sys.argv[1] == "--kotae":
        done = Path(sys.argv[2])
        rows = suck(done, cases)
    else:
        tmp = Path(tempfile.mkdtemp(prefix="kansu_"))
        src = tmp / "kansu_moto.xlsx"
        build_xlsx(cases, src)
        done = compute(src, tmp)
        rows = suck(done, cases)
        # 2010年以降の関数は xlsx の中では `_xlfn.` を頭に付けて書く決まり。
        # 付けずに #NAME? になった式へ機械的に付けて、もう1周だけ試す
        nokori = [
            (i, (name, formula.replace(name + "(", "_xlfn." + name + "(")))
            for i, ((name, formula), row) in enumerate(zip(cases, rows))
            if row.endswith("\t#NAME?")
        ]
        if nokori:
            src2 = tmp / "kansu_xlfn.xlsx"
            build_xlsx([c for _, c in nokori], src2)
            done2 = compute(src2, tmp)
            for (i, (name, _)), row2 in zip(nokori, suck(done2, [c for _, c in nokori])):
                v = row2.split("\t")[2]
                if v != "#NAME?":
                    # 表に載せる式は元の書き方のまま(答えだけ差し替え)
                    rows[i] = f"{name}\t{cases[i][1]}\t{v}"
    KOTAE.write_text(
        "# 関数の答えの表(機械が計算した物 — 手で直さない)。\n"
        "# 作り直し: python3 test/kansu_oracle.py。正はいま LibreOffice。\n"
        + "\n".join(rows)
        + "\n",
        encoding="utf-8",
    )
    print(f"{KOTAE.name}: {len(rows)} 件")


if __name__ == "__main__":
    main()
