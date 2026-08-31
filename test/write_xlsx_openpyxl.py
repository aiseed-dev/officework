#!/usr/bin/env python3
"""openpyxl の「書く」機能を、実在しそうな1ページの文書5種類に詰めて作る。

    .venv/bin/python test/write_xlsx_openpyxl.py [出力先フォルダ(既定 test/out)]

出来上がり(それぞれ1枚・A4 1ページに収まる形):
  見積書.xlsx        式・通貨の形式・結合・罫線・記入欄だけ開けた保護・
                     名前の定義・リンク・印刷の設定・ヘッダーフッター
  月次売上報告.xlsx  グラフ(縦棒+折れ線・円)・条件付き書式(バー/濃淡/赤字)・
                     %の形式・固定・テーブル・画像・枠線を消す
  出勤簿.xlsx        日付と時刻の形式・入力規則(リスト/時刻)・数式の条件で
                     土日に色・コメント・タイトル行の印刷・改ページ
  棚卸表.xlsx        文字列の形式(頭の0)・重複の色・オートフィルター・
                     グループ化と小計・別シート参照・隠しシート2種・配列式・
                     シート見出しの色
  申込書.xlsx        記入用紙の様式 — 面の結合・縦書き・回転・縮小・字下げ・
                     グラデーション・リッチテキスト・名前つきスタイル・
                     用紙に合わせる

同じ5枚を officework エンジンで作るのは test/write_xlsx_officework.py。
突き合わせは目で見る(発注者 2026-08-26)。
"""
import sys
import datetime as dt
from pathlib import Path

import openpyxl
from openpyxl.styles import (Font, PatternFill, GradientFill, Border, Side,
                             Alignment, Protection, NamedStyle)
from openpyxl.formatting.rule import (CellIsRule, ColorScaleRule, DataBarRule,
                                      FormulaRule, Rule)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.pagebreak import Break
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent / "out"
OUT.mkdir(parents=True, exist_ok=True)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
UNDER = Border(bottom=THIN)


def a4(ws, landscape=False):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    if landscape:
        ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


# ======================================================================
# 1. 見積書
# ======================================================================
def quote():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積書"
    for col, w in zip("ABCDEF", (4, 22, 8, 6, 12, 14)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    ws["A1"] = "御 見 積 書"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["E2"] = dt.date(2026, 8, 26)
    ws["E2"].number_format = "yyyy年m月d日"
    ws["E3"] = "見積番号"
    ws["F3"] = "Q-2026-0826"
    ws["F3"].number_format = "@"

    ws.merge_cells("A5:C5")
    ws["A5"] = "サンプル商事株式会社 御中"
    ws["A5"].font = Font(size=12, bold=True, underline="single")

    ws["E5"] = "aiseed"
    ws["E5"].hyperlink = "https://github.com/aiseed-dev/officework"
    ws["E5"].font = Font(color="0563C1", underline="single")
    ws["E6"] = "担当: 営業部"

    # 名前つきスタイルと名前の定義
    head = NamedStyle(name="表の見出し", font=Font(bold=True, color="FFFFFF"),
                      fill=PatternFill("solid", fgColor="1B6E3C"),
                      alignment=Alignment(horizontal="center"))
    wb.add_named_style(head)
    ws["F8"] = 0.1
    ws["F8"].number_format = "0%"
    wb.defined_names["税率"] = DefinedName("税率", attr_text="見積書!$F$8")
    ws["E8"] = "消費税率"

    # 明細の表
    for j, label in enumerate(["No.", "品名", "数量", "単位", "単価", "金額"], start=1):
        c = ws.cell(10, j, label)
        c.style = "表の見出し"
        c.border = BOX
    items = [("玄関ドア 親子 断熱", 1, "組", 458000),
             ("採光サイドパネル", 1, "枚", 76000),
             ("電気錠セット", 1, "式", 92000)]
    for i, (name, qty, unit, price) in enumerate(items, start=11):
        ws.cell(i, 1, i - 10).alignment = Alignment(horizontal="center")
        ws.cell(i, 2, name)
        ws.cell(i, 3, qty).alignment = Alignment(horizontal="right")
        ws.cell(i, 4, unit).alignment = Alignment(horizontal="center")
        p = ws.cell(i, 5, price)
        p.number_format = "#,##0"
        m = ws.cell(i, 6, f"=C{i}*E{i}")
        m.number_format = "#,##0"
        for j in range(1, 7):
            ws.cell(i, j).border = BOX
    ws["E15"] = "小計"
    ws["F15"] = "=SUM(F11:F13)"
    ws["E16"] = "消費税"
    ws["F16"] = "=F15*税率"
    ws["E17"] = "合計"
    ws["E17"].font = Font(bold=True)
    ws["F17"] = "=F15+F16"
    ws["F17"].font = Font(bold=True)
    for r in (15, 16, 17):
        ws.cell(r, 6).number_format = '"¥"#,##0'
        ws.cell(r, 5).border = UNDER
        ws.cell(r, 6).border = UNDER

    ws.merge_cells("A19:F19")
    ws["A19"] = "備考: 記入欄(黄色)だけ書き替えられます。ほかは保護しています。"
    ws["B20"] = "納期のご希望"
    memo = ws["C20"]
    memo.fill = PatternFill("solid", fgColor="FFF2CC")
    memo.protection = Protection(locked=False)
    memo.border = BOX

    ws.protection.sheet = True
    a4(ws)
    ws.oddFooter.center.text = "この見積の有効期限は発行から30日です"
    wb.properties.title = "御見積書"
    wb.properties.creator = "write_all_openpyxl"
    wb.save(OUT / "見積書.xlsx")


# ======================================================================
# 2. 月次売上報告
# ======================================================================
def monthly_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "8月"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    for col, w in zip("ABCDE", (12, 10, 10, 10, 10)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    ws["A1"] = "月次売上報告(2026年8月)"
    ws["A1"].font = Font(size=14, bold=True)

    logo = Path(__file__).resolve().parent.parent / "packaging/icons/hicolor/128x128/officework.png"
    if logo.exists():
        from openpyxl.drawing.image import Image
        ws.add_image(Image(str(logo)), "E2")

    header = ["支店", "目標", "実績", "達成率", "前月比"]
    data = [["本店", 5000, 5420], ["駅前店", 3000, 2710],
            ["西店", 2000, 2150], ["北店", 1500, 1180]]
    for j, h in enumerate(header, start=1):
        ws.cell(3, j, h)
    for i, (name, plan, act) in enumerate(data, start=4):
        ws.cell(i, 1, name)
        ws.cell(i, 2, plan).number_format = "#,##0"
        ws.cell(i, 3, act).number_format = "#,##0"
        r = ws.cell(i, 4, f"=C{i}/B{i}")
        r.number_format = "0.0%"
        d = ws.cell(i, 5, (0.06, -0.04, 0.11, -0.12)[i - 4])
        d.number_format = "+0.0%;-0.0%"
    tbl = Table(displayName="Sales", ref="A3:E7")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    # 条件付き書式: 実績にバー・達成率に濃淡・前月比のマイナスは赤字
    ws.conditional_formatting.add("C4:C7", DataBarRule(start_type="min", end_type="max",
                                                       color="638EC6"))
    ws.conditional_formatting.add("D4:D7", ColorScaleRule(start_type="min", start_color="FFFFFF",
                                                          end_type="max", end_color="63BE7B"))
    ws.conditional_formatting.add("E4:E7", CellIsRule(operator="lessThan", formula=["0"],
                                                      font=Font(color="9C0006")))

    data_ref = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=7)
    cats = Reference(ws, min_col=1, min_row=4, max_row=7)
    bar = BarChart()
    bar.title = "目標と実績"
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)
    line = LineChart()
    line.add_data(Reference(ws, min_col=4, max_col=4, min_row=3, max_row=7),
                  titles_from_data=True)
    bar += line
    ws.add_chart(bar, "A9")
    pie = PieChart()
    pie.title = "実績の内訳"
    pie.add_data(Reference(ws, min_col=3, min_row=3, max_row=7), titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "F9")

    a4(ws, landscape=True)
    wb.properties.title = "月次売上報告"
    wb.save(OUT / "月次売上報告.xlsx")


# ======================================================================
# 3. 出勤簿
# ======================================================================
def attendance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出勤簿"
    for col, w in zip("ABCDEF", (12, 8, 10, 10, 10, 16)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    ws["A1"] = "出勤簿(2026年8月・第4週)"
    ws["A1"].font = Font(size=13, bold=True)

    for j, h in enumerate(["日付", "区分", "出勤", "退勤", "実働", "備考"], start=1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
        c.border = BOX

    kinds = DataValidation(type="list", formula1='"出勤,休暇,半休,出張"', allow_blank=True)
    kinds.promptTitle = "区分"
    kinds.prompt = "一覧から選びます"
    kinds.errorTitle = "区分が違います"
    kinds.error = "出勤・休暇・半休・出張のどれかにしてください"
    ws.add_data_validation(kinds)
    hours = DataValidation(type="time", operator="between",
                           formula1="TIME(6,0,0)", formula2="TIME(23,0,0)")
    ws.add_data_validation(hours)

    base = dt.date(2026, 8, 24)
    for i in range(7):
        r = 4 + i
        d = ws.cell(r, 1, base + dt.timedelta(days=i))
        d.number_format = "m月d日(aaa)"
        ws.cell(r, 2, "出勤" if i < 5 else "休暇")
        kinds.add(ws.cell(r, 2))
        if i < 5:
            s = ws.cell(r, 3, dt.time(9, 0))
            e = ws.cell(r, 4, dt.time(17, 45 if i != 2 else 30))
            s.number_format = e.number_format = "h:mm"
            hours.add(s)
            hours.add(e)
            w = ws.cell(r, 5, f"=D{r}-C{r}-TIME(1,0,0)")
            w.number_format = "[h]:mm"
        for j in range(1, 7):
            ws.cell(r, j).border = BOX
    ws["F6"] = "客先で採寸"
    ws["F6"].comment = Comment("サンプル商事の現場です", "上長")

    # 土日の行に色(数式の条件)
    ws.conditional_formatting.add("A4:F10",
        FormulaRule(formula=['WEEKDAY($A4,2)>=6'],
                    fill=PatternFill("solid", fgColor="FCE4EC")))

    ws["D12"] = "実働の合計"
    t = ws["E12"]
    t.value = "=SUM(E4:E10)"
    t.number_format = "[h]:mm"
    t.font = Font(bold=True)

    ws.print_title_rows = "1:3"
    ws.row_breaks.append(Break(id=12))
    a4(ws)
    wb.properties.title = "出勤簿"
    wb.save(OUT / "出勤簿.xlsx")


# ======================================================================
# 4. 棚卸表
# ======================================================================
def inventory():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "棚卸"
    ws.sheet_properties.tabColor = "C00000"
    for col, w in zip("ABCDE", (10, 22, 8, 10, 12)):
        ws.column_dimensions[col].width = w

    prev = wb.create_sheet("前月")
    prev["A1"] = "品番"
    prev["B1"] = "数"
    for i, (code, n) in enumerate([("00123", 40), ("00456", 12), ("00789", 5)], start=2):
        c = prev.cell(i, 1, code)
        c.number_format = "@"
        prev.cell(i, 2, n)
    prev.sheet_state = "hidden"
    conf = wb.create_sheet("設定")
    conf["A1"] = "棚卸の基準日"
    conf["B1"] = dt.date(2026, 8, 31)
    conf.sheet_state = "veryHidden"

    ws.merge_cells("A1:E1")
    ws["A1"] = "棚卸表(倉庫A)"
    ws["A1"].font = Font(size=13, bold=True)

    for j, h in enumerate(["品番", "品名", "実数", "前月", "差"], start=1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True)
        c.border = BOX
    rows = [("00123", "玄関ドア用 丁番", 38), ("00456", "サイドパネル ガラス", 12),
            ("00789", "電気錠 制御基板", 4), ("00123", "玄関ドア用 丁番(再掲の誤り)", 2)]
    for i, (code, name, n) in enumerate(rows, start=4):
        c = ws.cell(i, 1, code)
        c.number_format = "@"          # 頭の 0 を守る
        ws.cell(i, 2, name)
        ws.cell(i, 3, n)
        ws.cell(i, 4, f'=IFERROR(VLOOKUP(A{i},前月!A:B,2,FALSE),0)')
        ws.cell(i, 5, f"=C{i}-D{i}")
        for j in range(1, 6):
            ws.cell(i, j).border = BOX
    # 品番の重複に色
    dup = Rule(type="duplicateValues",
               dxf=DifferentialStyle(fill=PatternFill(start_color="FFC7CE", fill_type="solid")))
    ws.conditional_formatting.add("A4:A7", dup)

    ws.auto_filter.ref = "A3:E7"
    # 行のグループ化と配列式の検算
    for r in (5, 6):
        ws.row_dimensions[r].outlineLevel = 1
    ws["C9"] = "実数の計"
    ws["D9"] = ArrayFormula("D9", "=SUM(C4:C7*1)")
    ws["A11"] = "基準日"
    ws["B11"] = "=設定!B1"
    ws["B11"].number_format = "yyyy/m/d"
    ws["A12"] = "前月の一覧へ"
    ws["A12"].hyperlink = "#前月!A1"
    ws["A12"].font = Font(color="0563C1", underline="single")

    a4(ws)
    wb.properties.title = "棚卸表"
    wb.save(OUT / "棚卸表.xlsx")


# ======================================================================
# 5. 申込書(記入用紙の様式)
# ======================================================================
def application_form():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "申込書"
    for col, w in zip("ABCDEFGH", (3, 4, 14, 14, 14, 14, 6, 3)):
        ws.column_dimensions[col].width = w
    ws.sheet_view.zoomScale = 110

    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = "会 員 申 込 書"
    t.font = Font(size=16, bold=True, color="FFFFFF")
    t.fill = GradientFill(stop=("1B6E3C", "63BE7B"))
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    note = NamedStyle(name="記入の案内", font=Font(size=9, color="808080"))
    wb.add_named_style(note)
    ws.merge_cells("B3:G3")
    ws["B3"] = CellRichText("太枠の中だけ ", TextBlock(InlineFont(b=True), "黒のボールペン"),
                            " でご記入ください(", TextBlock(InlineFont(color="FF0000"), "必須"),
                            " は空欄にできません)")
    ws["B3"].style = "記入の案内"

    # 縦書きの側柱
    ws.merge_cells("B5:B12")
    side = ws["B5"]
    side.value = "申込者"
    side.alignment = Alignment(horizontal="center", vertical="center", text_rotation=255)
    side.border = Border(left=MEDIUM, top=MEDIUM, bottom=MEDIUM, right=THIN)
    side.fill = PatternFill("solid", fgColor="EFEFEF")

    def field(row, label, span=("D", "F"), required=False):
        ws.cell(row, 3, label + ("(必須)" if required else "")).font = \
            Font(size=10, bold=required, color="C00000" if required else "000000")
        ws.cell(row, 3).alignment = Alignment(indent=1)
        ws.merge_cells(f"{span[0]}{row}:{span[1]}{row}")
        c = ws.cell(row, 4)
        c.protection = Protection(locked=False)
        c.border = BOX
        return c

    field(5, "ふりがな")
    field(6, "お名前", required=True)
    field(7, "生年月日", required=True).number_format = "yyyy年m月d日"
    field(8, "電話番号", required=True).number_format = "@"
    field(9, "メール")
    addr = field(10, "ご住所", required=True)
    addr.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[10].height = 30
    field(11, "ご紹介者").alignment = Alignment(shrink_to_fit=True)
    plan = field(12, "コース", span=("D", "D"), required=True)
    dv = DataValidation(type="list", formula1='"A(月2回),B(月4回),家族"')
    ws.add_data_validation(dv)
    dv.add(plan)

    # 外枠を太く
    for r in range(5, 13):
        ws.cell(r, 7).border = Border(right=MEDIUM)
    for col in range(2, 8):
        ws.cell(4, col).border = Border(bottom=MEDIUM)
        ws.cell(13, col).border = Border(top=MEDIUM)

    ws.merge_cells("C15:F15")
    ws["C15"] = "受付印"
    ws["C15"].alignment = Alignment(horizontal="right", vertical="top")
    ws.merge_cells("F16:F18")
    stamp = ws["F16"]
    stamp.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN,
                          diagonal=Side(style="hair"), diagonalDown=True)
    ws.cell(16, 3, "斜線の枠は事務局が使います").style = "記入の案内"

    ws.protection.sheet = True
    a4(ws)
    ws.oddHeader.right.text = "様式1"
    wb.properties.title = "会員申込書"
    wb.save(OUT / "申込書.xlsx")


if __name__ == "__main__":
    quote()
    monthly_report()
    attendance()
    inventory()
    application_form()
    print("5枚書けた:", ", ".join(
        f"{n}.xlsx" for n in ("見積書", "月次売上報告", "出勤簿", "棚卸表", "申込書")), "→", OUT)
