#!/usr/bin/env python3
"""突き合わせ用の xlsx を、**うちでもgenofficeでもない書き手**で作る。

`pyoffice_diff.py` は「向こうの答え」を正解表にするが、**両方が同じように
間違えていると差が出ない**。だから種は第三者に書かせる:

1. **openpyxl** が書く(書き手その1)。Excel とも LibreOffice とも違う XML を吐く
2. それを **LibreOffice** で開いて xlsx に保存し直す(書き手その2)。
   LibreOffice の書き出しは実物の世界で二番目に多い

**狙い撃ちで作る。** 統計の表は値と結合には強いが、条件付き書式・入力規則・
配列数式・リッチテキストはまず入っていない(2026-08-09 第1便で分かった)。
穴だと分かっている所を突く1枚を作るほうが、実物を10枚足すより速い。

    python3 tools/corpus_make.py                  # ~/xlsx-corpus/ へ
    python3 tools/corpus_make.py --out DIR
    python3 tools/corpus_make.py --no-lo          # LibreOffice の焼き直しを飛ばす

**現物は repo に置かない**(docs/corpus.ja.md)。この道具があれば作り直せる。
"""

import argparse
import os
import pathlib
import shutil
import subprocess
import sys

try:
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        FormulaRule,
        IconSetRule,
        Rule,
    )
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:
    sys.exit("openpyxl が要ります: pip install openpyxl")


def cond(out):
    """条件付き書式の9種を1枚に。**うちのモデルが持つ種類を全部踏む**。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "条件付き書式"
    ws["A1"] = "値"
    for i, v in enumerate([5, 12, 40, 88, 3, 12, 61, 7, 95, 40], start=2):
        ws.cell(i, 1, v)
        ws.cell(i, 2, v * 2)
        ws.cell(i, 3, f"品目{i % 4}")
    red = PatternFill("solid", start_color="FFC7CE")
    ws.conditional_formatting.add(
        "A2:A11", CellIsRule(operator="greaterThan", formula=["50"], fill=red)
    )
    ws.conditional_formatting.add(
        "A2:A11", CellIsRule(operator="between", formula=["10", "45"], fill=red)
    )
    ws.conditional_formatting.add(
        "C2:C11", Rule(type="containsText", text="品目1", dxf=None, formula=['NOT(ISERROR(SEARCH("品目1",C2)))'])
    )
    ws.conditional_formatting.add("A2:A11", Rule(type="duplicateValues"))
    ws.conditional_formatting.add("A2:A11", Rule(type="top10", rank=3))
    ws.conditional_formatting.add("A2:A11", Rule(type="aboveAverage"))
    ws.conditional_formatting.add(
        "B2:B11", DataBarRule(start_type="min", end_type="max", color="638EC6")
    )
    ws.conditional_formatting.add(
        "B2:B11",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF", mid_type="percentile", mid_value=50,
            mid_color="FFFF00", end_type="max", end_color="FF0000",
        ),
    )
    ws.conditional_formatting.add(
        "A2:A11", IconSetRule("3Arrows", "percent", [0, 33, 67])
    )
    ws.conditional_formatting.add("C2:C11", FormulaRule(formula=["LEN(C2)>3"], fill=red))
    wb.save(out / "make_cond.xlsx")


def valid(out):
    """入力規則の種類ぜんぶ。**知らない種類も落とさず持ち越す**が効くか。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "入力規則"
    kinds = [
        ("list", None, ['"甲,乙,丙"'], "一覧から選ぶ"),
        ("whole", "between", ["1", "100"], "1〜100 の整数"),
        ("decimal", "greaterThan", ["0.5"], "0.5 より大きい小数"),
        ("textLength", "lessThanOrEqual", ["8"], "8文字まで"),
        ("date", "greaterThan", ["2020-01-01"], "2020年より後"),
        ("time", "between", ["9:00", "18:00"], "就業時間"),
        ("custom", None, ["=ISNUMBER(F1)"], "数だけ"),
    ]
    for i, (kind, op, fs, note) in enumerate(kinds):
        col = chr(ord("A") + i)
        dv = DataValidation(type=kind, operator=op, formula1=fs[0],
                            formula2=fs[1] if len(fs) > 1 else None, allow_blank=True)
        dv.error = f"{note} を入れてください"
        dv.errorTitle = "入力の誤り"
        dv.prompt = note
        dv.promptTitle = kind
        dv.showErrorMessage = True
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}20")
        ws[f"{col}1"] = kind
    wb.save(out / "make_valid.xlsx")


def rich(out):
    """リッチテキスト(セルの中で書式が変わる)と、ふりがなの無い日本語。

    **うちが持たないと分かっている物**。潰すのか断るのかを決めるための1枚。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "リッチテキスト"
    ws["A1"] = CellRichText(
        "普通の字と",
        TextBlock(InlineFont(b=True, color="FF0000"), "赤い太字"),
        "と",
        TextBlock(InlineFont(i=True, sz=16), "大きい斜体"),
    )
    ws["A2"] = CellRichText(
        TextBlock(InlineFont(u="single"), "下線だけ"), "のこり"
    )
    ws["A3"] = "書式の変わらない普通のセル"
    ws["A4"] = CellRichText(TextBlock(InlineFont(strike=True), "取り消し線"))
    wb.save(out / "make_rich.xlsx")


def arrays(out):
    """昔ながらの配列数式(CSE)と、いろいろな式。

    CSE は `<f t="array" ref="…">` で書かれ、**読めないと黙って違う値になる**。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "式"
    for i in range(1, 6):
        ws.cell(i, 1, i)
        ws.cell(i, 2, i * 3)
    ws["D1"] = ArrayFormula("D1", "=SUM(A1:A5*B1:B5)")
    ws["D3"] = ArrayFormula("D3:D7", "=A1:A5*2")
    ws["F1"] = "=SUM(A1:A5)"
    ws["F2"] = "=IF(F1>10,\"多い\",\"少ない\")"
    ws["F3"] = "=VLOOKUP(3,A1:B5,2,FALSE)"
    ws["F4"] = "=TEXT(TODAY(),\"yyyy/mm/dd\")"
    ws["F5"] = "=COUNTIF(A1:A5,\">2\")"
    ws["F6"] = "=A1/0"
    ws["F7"] = "=SUM(名前つき)"
    wb.defined_names.add.__self__  # noqa: B018  (下で名前を足す)
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names["名前つき"] = DefinedName("名前つき", attr_text="式!$A$1:$A$5")
    wb.save(out / "make_arrays.xlsx")


def many_sheets(out):
    """**シートが10枚以上**で、並びが部品の番号と揃わない帳面。

    2026-08-09 に踏んだ「シートの取り違え」の受け皿。うちが自分で書く xlsx は
    sheet1..9 しか作らないので、この形は自分の答案では出ない。
    """
    wb = Workbook()
    names = [f"表{n}" for n in (1, 2, 3, 10, 11, 12, 20, 21, 30, 31, 40, 47)]
    wb.active.title = names[0]
    for n in names[1:]:
        wb.create_sheet(n)
    for i, n in enumerate(names):
        ws = wb[n]
        # **どのシートかが中身で分かる**ようにする。取り違えたら一目で出る
        ws["A1"] = f"これは {n} です"
        ws["A2"] = i * 100
        ws["B2"] = f"=A2*2"
    # 並びを部品の番号と食い違わせる(Excel でシートを動かしたのと同じ形)
    wb.move_sheet(names[-1], offset=-6)
    wb.move_sheet(names[1], offset=5)
    wb.save(out / "make_manysheets.xlsx")


def furniture(out):
    """帳票の造作 — 固定枠・結合・列幅・アウトライン・非表示・コメント・リンク・保護。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "台帳"
    ws.freeze_panes = "C3"
    ws.merge_cells("A1:E1")
    ws["A1"] = "見出しは結合で作る(日本の帳票)"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    for r in range(2, 12):
        for c in range(1, 6):
            cell = ws.cell(r, c, (r - 1) * c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions["A"].width = 24.5
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["B"].outlineLevel = 1
    ws.row_dimensions[5].height = 33.0
    ws.row_dimensions[7].hidden = True
    ws.row_dimensions[8].outlineLevel = 2
    ws["A3"].comment = Comment("この欄は税抜き", "検査")
    ws["B3"].hyperlink = "https://example.com/手引き"
    ws["B3"].value = "手引き"
    ws["C3"].font = Font(name="游ゴシック", size=14, bold=True, color="1F4E79")
    ws.sheet_properties.tabColor = "FF9900"
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.print_title_rows = "1:2"
    ws.oddHeader.center.text = "&P / &N 頁"
    hidden = wb.create_sheet("隠しシート")
    hidden["A1"] = "見えないが生きている"
    hidden.sheet_state = "hidden"
    rtl = wb.create_sheet("右横書き")
    rtl.sheet_view.rightToLeft = True
    rtl["A1"] = "右から左へ"
    wb.save(out / "make_furniture.xlsx")


def formats(out):
    """表示形式 — 日付・時刻・通貨・パーセント・桁区切り・負の赤・ユーザー定義。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "表示形式"
    rows = [
        ("日付", 45000, "yyyy年m月d日"),
        ("時刻", 0.5410, "h:mm:ss"),
        ("通貨", 1234567, '"¥"#,##0'),
        ("パーセント", 0.1234, "0.00%"),
        ("桁区切り", 9876543.21, "#,##0.00"),
        ("負は赤", -4567, "#,##0;[赤]-#,##0"),
        ("指数", 0.000123, "0.00E+00"),
        ("分数", 0.75, "# ?/?"),
        ("空欄は消す", 0, "#,##0;-#,##0;"),
    ]
    for i, (label, v, fmt) in enumerate(rows, start=1):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, v)
        c.number_format = fmt
    wb.save(out / "make_formats.xlsx")


MAKERS = [cond, valid, rich, arrays, many_sheets, furniture, formats]


def bake_with_libreoffice(out, names):
    """LibreOffice に開かせて xlsx で保存し直す — **書き手その2**。"""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        print("LibreOffice が見つからないので焼き直しは飛ばす")
        return
    tmp = out / "_lo"
    tmp.mkdir(exist_ok=True)
    for n in names:
        r = subprocess.run(
            [lo, "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(out / n)],
            capture_output=True, text=True, timeout=180,
        )
        src = tmp / n
        if r.returncode != 0 or not src.exists():
            # **黙って飛ばさない。** 焼けなかった物は名前を出す
            print(f"  焼き直せなかった: {n} — {(r.stderr or r.stdout).strip()[:120]}")
            continue
        dst = out / n.replace("make_", "lo_")
        shutil.move(str(src), str(dst))
        print(f"  {dst.name}")
    shutil.rmtree(tmp, ignore_errors=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.expanduser("~/xlsx-corpus"))
    ap.add_argument("--no-lo", action="store_true", help="LibreOffice の焼き直しを飛ばす")
    a = ap.parse_args()
    out = pathlib.Path(a.out)
    out.mkdir(parents=True, exist_ok=True)

    made = []
    for m in MAKERS:
        m(out)
        n = f"make_{m.__name__.replace('many_sheets', 'manysheets')}.xlsx"
        made.append(n)
        print(f"  {n}")
    if not a.no_lo:
        print("LibreOffice で焼き直す(書き手その2):")
        bake_with_libreoffice(out, made)
    print(f"\n{out} に置いた")


if __name__ == "__main__":
    main()
